#!/usr/bin/env python3
"""
run_experiments.py
==================
Executa múltiplas configurações do solver, coleta custos e tempos,
e gera um arquivo Excel (.xlsx) onde cada aba corresponde a um teste.

Uso rápido
----------
    python run_experiments.py                         # usa EXPERIMENT_SUITES padrão
    python run_experiments.py --binary ./build/optHirrygated
    python run_experiments.py --runs 30 --output resultados.xlsx
    python run_experiments.py --help
"""

import argparse
import subprocess
import statistics
import sys
import time
import os
from pathlib import Path
from dataclasses import dataclass, field
from typing import Optional

# ─────────────────────────────────────────────────────────────────────────────
#  CONFIGURAÇÃO DOS EXPERIMENTOS
#  Edite esta lista para adicionar/remover suítes de testes.
# ─────────────────────────────────────────────────────────────────────────────
EXPERIMENT_SUITES = [
    # ── Apenas construção ────────────────────────────────────────────────────
    {
        "name": "ForwardA_Apenas",
        "constructive": "forwardA",
        "refinement": "none",
        "metaheuristic": "none",
    },
    {
        "name": "ForwardB_Apenas",
        "constructive": "forwardB",
        "refinement": "none",
        "metaheuristic": "none",
    },
    {
        "name": "ForwardC_Apenas",
        "constructive": "forwardC",
        "refinement": "none",
        "metaheuristic": "none",
    },
    {
        "name": "Backward_Apenas",
        "constructive": "backward",
        "refinement": "none",
        "metaheuristic": "none",
    },
    {
        "name": "Lookahead_Apenas",
        "constructive": "lookahead",
        "refinement": "none",
        "metaheuristic": "none",
        "lookahead_depth": 3,
    },
    {
        "name": "LookaheadMemoization_Apenas",
        "constructive": "lookaheadMemoization",
        "refinement": "none",
        "metaheuristic": "none",
        "lookahead_depth": 5,
    },

    # ── Construção + VND ─────────────────────────────────────────────────────
    {
        "name": "ForwardA_VND",
        "constructive": "forwardA",
        "refinement": "vnd",
        "metaheuristic": "none",
    },
    {
        "name": "Backward_VND",
        "constructive": "backward",
        "refinement": "vnd",
        "metaheuristic": "none",
    },
    {
        "name": "Lookahead_VND",
        "constructive": "lookahead",
        "refinement": "vnd",
        "metaheuristic": "none",
        "lookahead_depth": 3,
    },
    {
        "name": "LookaheadMemoization_VND",
        "constructive": "lookaheadMemoization",
        "refinement": "vnd",
        "metaheuristic": "none",
        "lookahead_depth": 5,
    },

    # ── Construção + RVND ────────────────────────────────────────────────────
    {
        "name": "ForwardA_RVND",
        "constructive": "forwardA",
        "refinement": "rvnd",
        "metaheuristic": "none",
    },
    {
        "name": "Backward_RVND",
        "constructive": "backward",
        "refinement": "rvnd",
        "metaheuristic": "none",
    },
    {
        "name": "Lookahead_RVND",
        "constructive": "lookahead",
        "refinement": "rvnd",
        "metaheuristic": "none",
        "lookahead_depth": 3,
    },
    {
        "name": "LookaheadMemoization_RVND",
        "constructive": "lookaheadMemoization",
        "refinement": "rvnd",
        "metaheuristic": "none",
        "lookahead_depth": 5,
    },

    # ── ForwardA + VND + Metaheurísticas ─────────────────────────────────────
    {
        "name": "ForwardA_VND_MCTS",
        "constructive": "forwardA",
        "refinement": "vnd",
        "metaheuristic": "mcts",
    },
    {
        "name": "ForwardA_VND_ILS",
        "constructive": "forwardA",
        "refinement": "vnd",
        "metaheuristic": "ils",
        "ils_iterations": 100,
        "ils_perturb": 5,
    },
    {
        "name": "ForwardA_VND_PSO",
        "constructive": "forwardA",
        "refinement": "vnd",
        "metaheuristic": "pso",
    },

    # ── ForwardA + RVND + Metaheurísticas ────────────────────────────────────
    {
        "name": "ForwardA_RVND_MCTS",
        "constructive": "forwardA",
        "refinement": "rvnd",
        "metaheuristic": "mcts",
    },
    {
        "name": "ForwardA_RVND_ILS",
        "constructive": "forwardA",
        "refinement": "rvnd",
        "metaheuristic": "ils",
        "ils_iterations": 100,
        "ils_perturb": 5,
    },
    {
        "name": "ForwardA_RVND_PSO",
        "constructive": "forwardA",
        "refinement": "rvnd",
        "metaheuristic": "pso",
    },

    # ── Lookahead + VND + Metaheurísticas ────────────────────────────────────
    {
        "name": "Lookahead_VND_MCTS",
        "constructive": "lookahead",
        "refinement": "vnd",
        "metaheuristic": "mcts",
        "lookahead_depth": 3,
    },
    {
        "name": "Lookahead_VND_ILS",
        "constructive": "lookahead",
        "refinement": "vnd",
        "metaheuristic": "ils",
        "ils_iterations": 100,
        "ils_perturb": 5,
        "lookahead_depth": 3,
    },
    {
        "name": "Lookahead_VND_PSO",
        "constructive": "lookahead",
        "refinement": "vnd",
        "metaheuristic": "pso",
        "lookahead_depth": 3,
    },

    # ── Lookahead + RVND + Metaheurísticas ───────────────────────────────────
    {
        "name": "Lookahead_RVND_MCTS",
        "constructive": "lookahead",
        "refinement": "rvnd",
        "metaheuristic": "mcts",
        "lookahead_depth": 3,
    },
    {
        "name": "Lookahead_RVND_ILS",
        "constructive": "lookahead",
        "refinement": "rvnd",
        "metaheuristic": "ils",
        "ils_iterations": 100,
        "ils_perturb": 5,
        "lookahead_depth": 3,
    },
    {
        "name": "Lookahead_RVND_PSO",
        "constructive": "lookahead",
        "refinement": "rvnd",
        "metaheuristic": "pso",
        "lookahead_depth": 3,
    },
]
# ─────────────────────────────────────────────────────────────────────────────
#  Argument parsing
# ─────────────────────────────────────────────────────────────────────────────

def build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        description="Roda múltiplas suítes de experimentos e gera relatório Excel.",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    p.add_argument("-n", "--runs", type=int, default=10,
                   help="Número de execuções por suíte")
    p.add_argument("--binary", type=str, default="./build/optHirrygated",
                   help="Caminho para o binário compilado")
    p.add_argument("--datasource", type=str, default="./datasource/planilha.xlsx",
                   help="Caminho para o arquivo .xlsx de dados")
    p.add_argument("--timeout", type=float, default=3000.0,
                   help="Timeout por execução em segundos")
    p.add_argument("--output", type=str, default="experimentos.xlsx",
                   help="Arquivo Excel de saída")
    p.add_argument("--verbose", action="store_true",
                   help="Mostra a saída bruta de cada execução")
    p.add_argument("--suites", type=str, default=None,
                   help="Nomes das suítes a executar (separados por vírgula). "
                        "Omitir para executar todas.")
    return p


# ─────────────────────────────────────────────────────────────────────────────
#  Build solver command
# ─────────────────────────────────────────────────────────────────────────────

def build_cmd(binary: str, datasource: str, suite: dict) -> list[str]:
    cmd = [
        binary,
        "--constructive",    suite.get("constructive", "forward"),
        "--lookahead-depth", str(suite.get("lookahead_depth", 2)),
        "--refinement",      suite.get("refinement", "none"),
        "--metaheuristic",   suite.get("metaheuristic", "none"),
        "--ils-iterations",  str(suite.get("ils_iterations", 100)),
        "--ils-perturb",     str(suite.get("ils_perturb", 5)),
        "--datasource",      datasource,
        "--csv",
    ]
    return cmd


# ─────────────────────────────────────────────────────────────────────────────
#  Runner
# ─────────────────────────────────────────────────────────────────────────────

def run_once(cmd: list[str], timeout: float, verbose: bool) -> Optional[dict]:
    """
    Executa o solver uma vez.
    Retorna dict com campos + elapsed_seconds, ou None em caso de erro.
    """
    t0 = time.perf_counter()
    try:
        result = subprocess.run(
            cmd,
            capture_output=True,
            text=True,
            timeout=timeout,
        )
        elapsed = time.perf_counter() - t0
    except subprocess.TimeoutExpired:
        print("  [TIMEOUT]", flush=True)
        return None
    except FileNotFoundError:
        print(f"\n[ERRO] Binário não encontrado: {cmd[0]}", file=sys.stderr)
        sys.exit(1)

    if verbose:
        print("  STDOUT:", result.stdout.strip())
        if result.stderr.strip():
            print("  STDERR:", result.stderr.strip())

    if result.returncode != 0:
        print(f"  [ERRO] returncode={result.returncode}", flush=True)
        return None

    lines = result.stdout.strip().splitlines()
    if not lines:
        print("  [ERRO] Sem saída do solver", flush=True)
        return None

    line = lines[-1]
    parts = line.split(",")
    if len(parts) != 5:
        print(f"  [ERRO] Output inesperado: {line!r}", flush=True)
        return None

    constructive, refinement, metaheuristic, valid_str, cost_str = parts
    try:
        cost = float(cost_str)
    except ValueError:
        print(f"  [ERRO] Custo inválido: {cost_str!r}", flush=True)
        return None

    return {
        "constructive":   constructive,
        "refinement":     refinement,
        "metaheuristic":  metaheuristic,
        "valid":          valid_str.strip() == "1",
        "cost":           cost,
        "elapsed":        elapsed,
    }


def run_suite(suite: dict, runs: int, binary: str, datasource: str,
              timeout: float, verbose: bool) -> tuple[list[dict], int]:
    """Executa uma suíte completa. Retorna (rows, errors)."""
    cmd = build_cmd(binary, datasource, suite)
    rows = []
    errors = 0
    sep = "─" * 56

    print(f"\n{sep}")
    print(f"  SUÍTE: {suite['name']}")
    print(f"    Construtiva    : {suite.get('constructive','forward')}")
    print(f"    Refinamento    : {suite.get('refinement','none')}")
    print(f"    Metaheurística : {suite.get('metaheuristic','none')}")
    print(sep)

    for i in range(1, runs + 1):
        print(f"  [{i:>3}/{runs}] ", end="", flush=True)
        data = run_once(cmd, timeout=timeout, verbose=verbose)
        if data is None:
            errors += 1
            print("FALHA")
            continue
        status = "OK" if data["valid"] else "INVÁLIDA"
        print(f"R$ {data['cost']:.4f}  |  {data['elapsed']:.2f}s  [{status}]")
        rows.append(data)

    return rows, errors


# ─────────────────────────────────────────────────────────────────────────────
#  Statistics
# ─────────────────────────────────────────────────────────────────────────────

def compute_stats(values: list[float]) -> dict:
    n = len(values)
    if n == 0:
        return {}
    mean   = statistics.mean(values)
    best   = min(values)
    worst  = max(values)
    std    = statistics.stdev(values) if n > 1 else 0.0
    median = statistics.median(values)
    cv     = (std / mean * 100) if mean != 0 else 0.0
    return dict(n=n, mean=mean, best=best, worst=worst,
                std=std, median=median, cv=cv)


# ─────────────────────────────────────────────────────────────────────────────
#  Excel export
# ─────────────────────────────────────────────────────────────────────────────

def _col_letter(n: int) -> str:
    """Converte número de coluna (1-based) para letra Excel."""
    result = ""
    while n:
        n, r = divmod(n - 1, 26)
        result = chr(65 + r) + result
    return result


def create_excel(output_path: str, suite_results: list[tuple[dict, list[dict], int]]) -> None:
    """
    suite_results: lista de (suite_config, rows, errors)
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import (Font, PatternFill, Alignment,
                                     Border, Side, numbers)
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("[ERRO] openpyxl não encontrado. Instale com: pip install openpyxl",
              file=sys.stderr)
        sys.exit(1)

    # ── Color palette ────────────────────────────────────────────────────────
    C_HEADER_BG  = "1F4E79"   # azul escuro
    C_HEADER_FG  = "FFFFFF"   # branco
    C_STATS_BG   = "D6E4F0"   # azul claro
    C_STATS_LABEL= "1F4E79"   # azul escuro
    C_BEST_BG    = "C6EFCE"   # verde claro
    C_WORST_BG   = "FFC7CE"   # vermelho claro
    C_ALT_ROW    = "EBF5FB"   # azul muito claro (linhas alternadas)
    C_INVALID_BG = "FFE599"   # amarelo (execução inválida)
    C_BORDER     = "BFC9CA"   # cinza claro

    thin_side = Side(style="thin", color=C_BORDER)
    thin_border = Border(left=thin_side, right=thin_side,
                         top=thin_side, bottom=thin_side)

    def hdr_style(cell, bg=C_HEADER_BG, fg=C_HEADER_FG):
        cell.font      = Font(bold=True, color=fg, name="Arial", size=10)
        cell.fill      = PatternFill("solid", start_color=bg)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = thin_border

    def stat_label(cell):
        cell.font      = Font(bold=True, color=C_STATS_LABEL, name="Arial", size=10)
        cell.fill      = PatternFill("solid", start_color=C_STATS_BG)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border    = thin_border

    def stat_value(cell, number_fmt='#,##0.0000'):
        cell.font      = Font(name="Arial", size=10)
        cell.fill      = PatternFill("solid", start_color=C_STATS_BG)
        cell.alignment = Alignment(horizontal="right", vertical="center")
        cell.number_format = number_fmt
        cell.border    = thin_border

    def data_cell(cell, align="right"):
        cell.font      = Font(name="Arial", size=10)
        cell.alignment = Alignment(horizontal=align, vertical="center")
        cell.border    = thin_border

    wb = Workbook()
    wb.remove(wb.active)  # remove sheet padrão

    # ── Aba de Resumo ────────────────────────────────────────────────────────
    summary_ws = wb.create_sheet("Resumo")

    # Título
    summary_ws.merge_cells("A1:I1")
    title_cell = summary_ws["A1"]
    title_cell.value     = "Resumo dos Experimentos – OptHirrygated"
    title_cell.font      = Font(bold=True, color=C_HEADER_FG, name="Arial", size=13)
    title_cell.fill      = PatternFill("solid", start_color=C_HEADER_BG)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    summary_ws.row_dimensions[1].height = 28

    summary_ws.merge_cells("A2:I2")
    import datetime
    summary_ws["A2"].value     = f"Gerado em {datetime.datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
    summary_ws["A2"].font      = Font(italic=True, color="666666", name="Arial", size=9)
    summary_ws["A2"].alignment = Alignment(horizontal="center")

    # Cabeçalho resumo
    sum_headers = ["Suíte", "Construtiva", "Refinamento", "Metaheurística",
                   "Execuções", "Melhor (R$)", "Pior (R$)", "Média (R$)", "Desv. Padrão (R$)"]
    for c, h in enumerate(sum_headers, 1):
        cell = summary_ws.cell(row=4, column=c, value=h)
        hdr_style(cell)
    summary_ws.row_dimensions[4].height = 20

    sum_col_widths = [22, 14, 14, 16, 10, 14, 14, 14, 16]
    for i, w in enumerate(sum_col_widths, 1):
        summary_ws.column_dimensions[get_column_letter(i)].width = w

    # ── Uma aba por suíte ────────────────────────────────────────────────────
    for suite_idx, (suite, rows, errors) in enumerate(suite_results):
        sheet_name = suite["name"][:31]  # Excel limita a 31 chars
        ws = wb.create_sheet(sheet_name)

        costs   = [r["cost"]    for r in rows]
        times   = [r["elapsed"] for r in rows]
        c_stats = compute_stats(costs)
        t_stats = compute_stats(times)

        best_cost  = min(costs)  if costs else None
        worst_cost = max(costs)  if costs else None

        # ── Título da aba ────────────────────────────────────────────────────
        ws.merge_cells("A1:H1")
        tc = ws["A1"]
        tc.value     = f"Experimento: {suite['name']}"
        tc.font      = Font(bold=True, color=C_HEADER_FG, name="Arial", size=12)
        tc.fill      = PatternFill("solid", start_color=C_HEADER_BG)
        tc.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 26

        # ── Bloco de configuração ────────────────────────────────────────────
        config_labels = [
            ("Construtiva",    suite.get("constructive", "forward")),
            ("Refinamento",    suite.get("refinement",   "none")),
            ("Metaheurística", suite.get("metaheuristic","none")),
        ]
        if suite.get("metaheuristic") == "ils":
            config_labels += [
                ("ILS Iterações", str(suite.get("ils_iterations", 100))),
                ("ILS Perturbação", str(suite.get("ils_perturb", 5))),
            ]
        if suite.get("constructive") == "lookahead":
            config_labels.insert(1, ("Profundidade Lookahead", str(suite.get("lookahead_depth", 2))))

        for i, (lbl, val) in enumerate(config_labels):
            r = 2 + i
            lc = ws.cell(row=r, column=1, value=lbl)
            vc = ws.cell(row=r, column=2, value=val)
            lc.font      = Font(bold=True, name="Arial", size=10)
            lc.alignment = Alignment(horizontal="left", vertical="center")
            lc.border    = thin_border
            vc.font      = Font(name="Arial", size=10)
            vc.alignment = Alignment(horizontal="left", vertical="center")
            vc.border    = thin_border

        cfg_rows = 2 + len(config_labels)

        # ── Cabeçalho da tabela de execuções ─────────────────────────────────
        data_start_row = cfg_rows + 2
        exec_headers = ["Execução", "Custo (R$)", "Tempo (s)", "Válida"]
        for c, h in enumerate(exec_headers, 1):
            cell = ws.cell(row=data_start_row, column=c, value=h)
            hdr_style(cell)
        ws.row_dimensions[data_start_row].height = 20

        # ── Linhas de execução ───────────────────────────────────────────────
        for i, row in enumerate(rows):
            r = data_start_row + 1 + i
            ws.row_dimensions[r].height = 18

            # Linha alternada
            row_bg = C_ALT_ROW if i % 2 == 1 else "FFFFFF"
            if not row["valid"]:
                row_bg = C_INVALID_BG

            for c in range(1, 5):
                cell = ws.cell(row=r, column=c)
                cell.fill   = PatternFill("solid", start_color=row_bg)
                cell.border = thin_border
                cell.font   = Font(name="Arial", size=10)

            ws.cell(row=r, column=1).value          = i + 1
            ws.cell(row=r, column=1).alignment      = Alignment(horizontal="center")

            cost_cell = ws.cell(row=r, column=2)
            cost_cell.value          = row["cost"]
            cost_cell.number_format  = 'R$ #,##0.0000'
            cost_cell.alignment      = Alignment(horizontal="right")

            # Destaque melhor/pior
            if best_cost is not None and abs(row["cost"] - best_cost) < 1e-9:
                cost_cell.fill = PatternFill("solid", start_color=C_BEST_BG)
            elif worst_cost is not None and abs(row["cost"] - worst_cost) < 1e-9:
                cost_cell.fill = PatternFill("solid", start_color=C_WORST_BG)

            time_cell = ws.cell(row=r, column=3)
            time_cell.value         = row["elapsed"]
            time_cell.number_format = '0.00"s"'
            time_cell.alignment     = Alignment(horizontal="right")

            valid_cell = ws.cell(row=r, column=4)
            valid_cell.value     = "Sim" if row["valid"] else "Não"
            valid_cell.alignment = Alignment(horizontal="center")

        last_data_row = data_start_row + len(rows)

        # ── Bloco de estatísticas de custo ───────────────────────────────────
        stats_start = last_data_row + 2
        ws.merge_cells(f"A{stats_start}:D{stats_start}")
        sh = ws[f"A{stats_start}"]
        sh.value     = "Estatísticas – Custo (R$)"
        sh.font      = Font(bold=True, color=C_HEADER_FG, name="Arial", size=11)
        sh.fill      = PatternFill("solid", start_color=C_HEADER_BG)
        sh.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[stats_start].height = 20

        cost_stat_rows = [
            ("Execuções",        c_stats.get("n", 0),     '0'),
            ("Falhas",           errors,                   '0'),
            ("Melhor (mín)",     c_stats.get("best"),      'R$ #,##0.0000'),
            ("Pior (máx)",       c_stats.get("worst"),     'R$ #,##0.0000'),
            ("Média",            c_stats.get("mean"),      'R$ #,##0.0000'),
            ("Mediana",          c_stats.get("median"),    'R$ #,##0.0000'),
            ("Desvio Padrão",    c_stats.get("std"),       'R$ #,##0.0000'),
            ("Coef. Variação",   c_stats.get("cv"),        '0.00"%"'),
        ]

        for j, (lbl, val, fmt) in enumerate(cost_stat_rows):
            r = stats_start + 1 + j
            lc = ws.cell(row=r, column=1, value=lbl)
            vc = ws.cell(row=r, column=2, value=val)
            ws.merge_cells(f"A{r}:A{r}")
            ws.merge_cells(f"B{r}:D{r}")
            stat_label(lc)
            stat_value(vc, fmt)

        # ── Bloco de estatísticas de tempo ───────────────────────────────────
        time_stats_start = stats_start + len(cost_stat_rows) + 2
        ws.merge_cells(f"A{time_stats_start}:D{time_stats_start}")
        th = ws[f"A{time_stats_start}"]
        th.value     = "Estatísticas – Tempo de Execução (s)"
        th.font      = Font(bold=True, color=C_HEADER_FG, name="Arial", size=11)
        th.fill      = PatternFill("solid", start_color=C_HEADER_BG)
        th.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[time_stats_start].height = 20

        time_stat_rows = [
            ("Tempo Mínimo",  t_stats.get("best"),   '0.00"s"'),
            ("Tempo Máximo",  t_stats.get("worst"),  '0.00"s"'),
            ("Tempo Médio",   t_stats.get("mean"),   '0.00"s"'),
            ("Desvio Padrão", t_stats.get("std"),    '0.00"s"'),
        ]

        for j, (lbl, val, fmt) in enumerate(time_stat_rows):
            r = time_stats_start + 1 + j
            lc = ws.cell(row=r, column=1, value=lbl)
            vc = ws.cell(row=r, column=2, value=val)
            ws.merge_cells(f"B{r}:D{r}")
            stat_label(lc)
            stat_value(vc, fmt)

        # ── Larguras das colunas ─────────────────────────────────────────────
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 14
        ws.column_dimensions["D"].width = 10

        # ── Linha de legenda ─────────────────────────────────────────────────
        legend_row = time_stats_start + len(time_stat_rows) + 2
        ws.merge_cells(f"A{legend_row}:D{legend_row}")
        lg = ws[f"A{legend_row}"]
        lg.value     = "🟢 Melhor custo   🔴 Pior custo   🟡 Execução inválida"
        lg.font      = Font(italic=True, color="444444", name="Arial", size=9)
        lg.alignment = Alignment(horizontal="left")

        # ── Adiciona linha no resumo ─────────────────────────────────────────
        sum_row = 4 + suite_idx + 1
        summary_ws.cell(row=sum_row, column=1, value=suite["name"])
        summary_ws.cell(row=sum_row, column=2, value=suite.get("constructive","forward"))
        summary_ws.cell(row=sum_row, column=3, value=suite.get("refinement","none"))
        summary_ws.cell(row=sum_row, column=4, value=suite.get("metaheuristic","none"))
        summary_ws.cell(row=sum_row, column=5, value=c_stats.get("n", 0))

        for col_idx, key in enumerate(["best","worst","mean","std"], 6):
            cell = summary_ws.cell(row=sum_row, column=col_idx,
                                   value=c_stats.get(key))
            cell.number_format = 'R$ #,##0.0000'
            cell.alignment     = Alignment(horizontal="right")

        # Estilo alternado no resumo
        row_bg = C_ALT_ROW if suite_idx % 2 == 1 else "FFFFFF"
        for col in range(1, 10):
            c = summary_ws.cell(row=sum_row, column=col)
            if c.fill.start_color.rgb in ("00000000", "FFFFFFFF", "00FFFFFF"):
                c.fill = PatternFill("solid", start_color=row_bg)
            c.font   = Font(name="Arial", size=10)
            c.border = thin_border
            c.alignment = Alignment(horizontal="center" if col in (1,2,3,4,5) else "right",
                                    vertical="center")

    # ── Salva ────────────────────────────────────────────────────────────────
    wb.save(output_path)
    print(f"\n  ✅  Relatório Excel salvo em: {output_path}")


# ─────────────────────────────────────────────────────────────────────────────
#  Main
# ─────────────────────────────────────────────────────────────────────────────

def main() -> None:
    parser = build_parser()
    args   = parser.parse_args()

    # Sanity check: binary
    if not Path(args.binary).exists():
        print(f"[AVISO] Binário não encontrado em '{args.binary}'.\n"
              f"        Compile o projeto e use --binary <caminho>.",
              file=sys.stderr)
        sys.exit(1)

    # Filtra suítes se --suites foi passado
    suites = EXPERIMENT_SUITES
    if args.suites:
        names = {s.strip() for s in args.suites.split(",")}
        suites = [s for s in suites if s["name"] in names]
        if not suites:
            print(f"[ERRO] Nenhuma suíte encontrada com os nomes: {args.suites}",
                  file=sys.stderr)
            sys.exit(1)

    print(f"\n{'═'*56}")
    print(f"  OptHirrygated – Runner de Experimentos")
    print(f"  Suítes   : {len(suites)}")
    print(f"  Execuções: {args.runs} por suíte")
    print(f"  Saída    : {args.output}")
    print(f"{'═'*56}")

    suite_results = []

    for suite in suites:
        rows, errors = run_suite(
            suite, args.runs,
            binary=args.binary,
            datasource=args.datasource,
            timeout=args.timeout,
            verbose=args.verbose,
        )

        costs = [r["cost"] for r in rows]
        if not costs:
            print(f"  ⚠  Suíte '{suite['name']}': nenhuma execução bem-sucedida.")
        else:
            s = compute_stats(costs)
            print(f"\n  Resumo '{suite['name']}':")
            print(f"    Melhor  : R$ {s['best']:.4f}")
            print(f"    Média   : R$ {s['mean']:.4f}")
            print(f"    Desvio  : R$ {s['std']:.4f}")
            if errors:
                print(f"    Falhas  : {errors}")

        suite_results.append((suite, rows, errors))

    # Gera Excel
    print(f"\n{'─'*56}")
    print("  Gerando relatório Excel…")
    create_excel(args.output, suite_results)


if __name__ == "__main__":
    main()